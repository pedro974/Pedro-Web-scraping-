from selenium import  webdriver
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
class google:
    def __init__(self,driver):
        self.driver = driver
        self.url = 'https://www.olx.com.br'
        self.search_b = 'q'
        self.pesquisar_boton = 'submitBtn'              
        self.titulo= 'OLXad-list-title' #class titulo
        self.local = 'OLXad-list-line-2' #classe local
        self.categoria = 'text detail-category' #class categoria
        self.preco = 'col-3' #class preco
        self.horario = 'col-4' #class horario
        self.next_pagina = 'item next'
    def navigate(self):
         self.driver.get(self.url)
    def search(self, word = 'None'):
        self.driver.find_element_by_name(self.search_b).send_keys(word)
        self.driver.find_element_by_class_name(self.pesquisar_boton).click()
    def printx(self):
        h= 1
        a =self.driver.find_elements_by_class_name(self.titulo)
        for g in a:
            h +=1
            gx = g.text
            ws.cell(row=h,column =2, value = gx)
            print(h)
        b =self.driver.find_elements_by_class_name(self.preco)
        hx = 1
        for g in b:
            hx +=1
            gx = g.text
            ws.cell(row=hx,column =3, value = gx)
            print(hx)
        c =self.driver.find_elements_by_class_name(self.local)
        hg = 1
        for g in c:
            hg +=1
            gx = g.text
            ws.cell(row=hg,column =4, value = gx)
            print(hg)
        hy = 1
        d =self.driver.find_elements_by_class_name(self.horario)
        for g in d:
            hy +=1
            gx = g.text
            ws.cell(row=hy,column =5, value = gx)
            print(hy)
        e =self.driver.find_elements_by_class_name(self.categoria)
        hh = 1
        for g in e:
            hh +=1
            gx = g.text
            ws.cell(row=hh,column =6, value = gx)
            print(hh)
        def pagina(self):
            self.driver.find_element_by_class_name(self.next_pagina).click()
       
        

        
        
        
                    
  
  
gg = webdriver.Chrome() #abre o driver
g = google(gg) #abre o google
g.navigate() #abre navegador
g.search('Celular') #c
g.printx
wb.save('olx.xlsx')
g.pagina()
g.printx()
wb.save('olx.xlsx')
gg.close()


"""for local in c:
#print(local.text)
for horario in d:
#print(horario.text)
for categoria in f:""
                            """
